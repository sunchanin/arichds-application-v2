"""Eager capture on a closed Billing Reading insert (decision 11, SPEC §3.6,
issue #22).

Wired into ``read_and_store_billing`` after the Transport Endpoint lock is
released and after the row commits — never inside the lock, never for the
Open Period. Gated on ``auto_capture`` (PDF) / ``billing_excel_export``
(xlsx alongside it), read through the process-wide LicenseService holder
(``licensing.current``) since this path has no request.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from fakes import FakeMeterState

from arichds.acquisition.billing import read_and_store_billing
from arichds.acquisition.drivers.base import BillingReading
from arichds.acquisition.locks import EndpointLocks
from arichds.config import Settings
from arichds.constants import SOURCE_DLMS
from arichds.db.app_settings import CAPTURE_DIR_KEY, DISPLAY_UNIT_SCALE_KEY, set_setting
from arichds.db.session import session_scope
from arichds.licensing.current import set_current_license_service
from arichds.licensing.service import LicenseState

pytestmark = pytest.mark.usefixtures("fake_meter")

NOW = datetime(2026, 8, 7, 6, 0, tzinfo=UTC)
#: Matches `make_device()`'s transport below — the Poller's lock key
#: (`Device.transport_endpoint`) for `{"kind": "net", "host": "127.0.0.1", "port": 4059}`.
ENDPOINT = "127.0.0.1:4059"

ENTRY_CLOSED = BillingReading(
    bill_date=datetime(2026, 7, 31, 17, 0, 0, tzinfo=UTC),
    source=SOURCE_DLMS,
    is_open=False,
    meter_serial="1232002893",
    import_active_kwh_total=198685.030,
)
ENTRY_OPEN = BillingReading(
    bill_date=datetime(2026, 8, 7, 4, 37, 58, tzinfo=UTC),
    source=SOURCE_DLMS,
    is_open=True,
    meter_serial="1232002893",
    import_active_kwh_total=200464.501,
)


class _StubLicenseService:
    """A minimal stand-in satisfying `feature_enabled`'s one call —
    `.current_state()` — without going through a real activation flow."""

    def __init__(self, features: list[str] | None) -> None:
        self._features = features

    def current_state(self) -> LicenseState:
        return LicenseState(state="active", reason=None, features=self._features)


@pytest.fixture
def license_features():
    def apply(features: list[str] | None) -> None:
        set_current_license_service(_StubLicenseService(features))

    yield apply
    set_current_license_service(None)


def make_device() -> int:
    from arichds.db.models import Device

    with session_scope() as session:
        device = Device(
            name="Main Incomer",
            brand="mitsu",
            model="smw110",
            site_name="Plant A",
            transport={"kind": "net", "host": "127.0.0.1", "port": 4059},
            password="hunter2",
        )
        session.add(device)
        session.flush()
        return device.id


@pytest.fixture
def device_id(migrated_db: Settings) -> int:
    return make_device()


@pytest.fixture
def capture_dir(tmp_path: Path) -> Path:
    """A directory distinct from `data_dir`'s tree — the `data_dir` fixture
    (through `migrated_db`) also lives under `tmp_path`, and pointing
    `capture_dir` at the same root would make `captured_files()` pick up the
    SQLite file and its WAL/SHM siblings."""
    target = tmp_path / "captures"
    target.mkdir()
    return target


def set_capture_dir(path: Path) -> None:
    with session_scope() as session:
        set_setting(session, CAPTURE_DIR_KEY, str(path))


def captured_files(path: Path) -> list[Path]:
    return sorted(p for p in path.rglob("*") if p.is_file())


class TestBothFeaturesOn:
    def test_pdf_and_xlsx_are_both_written(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing", "auto_capture", "billing_excel_export"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        files = captured_files(capture_dir)
        assert any(f.suffix == ".pdf" for f in files)
        assert any(f.suffix == ".xlsx" for f in files)


class TestDisplayUnitScaleReachesTheEagerCapture:
    def test_base_scale_is_applied_to_the_written_xlsx(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        """The setting is read fresh on every capture, not cached — proven end
        to end through the real write path, not by inspecting the call."""
        from openpyxl import load_workbook

        license_features(["billing", "auto_capture", "billing_excel_export"])
        set_capture_dir(capture_dir)
        with session_scope() as session:
            set_setting(session, DISPLAY_UNIT_SCALE_KEY, "base")
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        xlsx_path = next(f for f in captured_files(capture_dir) if f.suffix == ".xlsx")
        values = [cell.value for row_cells in load_workbook(xlsx_path).active.iter_rows() for cell in row_cells]
        assert "Import Active Wh Total" in values
        assert "Import Active kWh Total" not in values


class TestOnlyAutoCaptureOn:
    def test_only_the_pdf_is_written(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing", "auto_capture"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        files = captured_files(capture_dir)
        assert any(f.suffix == ".pdf" for f in files)
        assert not any(f.suffix == ".xlsx" for f in files)


class TestNeitherFeatureOn:
    def test_nothing_is_written(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing"])  # neither auto_capture nor billing_excel_export
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        assert captured_files(capture_dir) == []


class TestOpenPeriodNeverCaptured:
    def test_only_open_row_produces_no_capture(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing", "auto_capture", "billing_excel_export"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_OPEN]

        read_and_store_billing(device_id, now=NOW)

        assert captured_files(capture_dir) == []

    def test_a_mixed_buffer_captures_only_the_closed_row(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing", "auto_capture"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_OPEN, ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        files = captured_files(capture_dir)
        assert len(files) == 1
        assert "2026-07-31" in files[0].name  # the closed entry's bill_date, not the open one's


class TestNoCaptureDirConfigured:
    def test_nothing_is_written_and_the_read_still_succeeds(
        self, device_id: int, fake_meter: FakeMeterState, license_features
    ) -> None:
        license_features(["billing", "auto_capture", "billing_excel_export"])
        # capture_dir left at its default ("") — never configured.
        fake_meter.billing_rows = [ENTRY_CLOSED]

        result = read_and_store_billing(device_id, now=NOW)

        assert result.error is None
        assert result.stored == 1


class TestCaptureFailureNeverFailsTheRead:
    def test_a_capture_that_raises_is_logged_and_the_read_still_reports_success(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        import arichds.acquisition.billing as billing_module

        license_features(["billing", "auto_capture"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("capture blew up")

        monkeypatch.setattr(billing_module, "capture_reading", boom)

        with caplog.at_level("ERROR"):
            result = read_and_store_billing(device_id, now=NOW)

        assert result.error is None
        assert result.stored == 1
        assert any("capture" in record.message.lower() for record in caplog.records)


class TestNoLicenseServicePublished:
    def test_no_service_means_no_capture_and_no_crash(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path
    ) -> None:
        """`licensing.current` starts with nothing published — background code
        run outside the lifespan must treat that as "feature off", never crash."""
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        result = read_and_store_billing(device_id, now=NOW)

        assert result.error is None
        assert captured_files(capture_dir) == []


class TestCaptureRunsOutsideTheEndpointLock:
    def test_the_endpoint_lock_is_free_the_moment_capture_runs(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """SPEC §3.6: capture is eager but runs OUTSIDE the Transport Endpoint
        lock — a slow network share must never delay another meter on the
        same line, or make a Manual Read wait behind a capture write.

        Proven by trying to acquire the SAME lock object, for the SAME
        endpoint, from inside the (patched) capture path: `background()`
        only succeeds when nothing else holds the lock. If capture still ran
        *inside* the read's `with registry.get(endpoint).manual(...)` block,
        this acquisition would find the lock held and return `False`.
        """
        import arichds.acquisition.billing as billing_module

        license_features(["billing", "auto_capture"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        locks = EndpointLocks()
        acquired_during_capture: list[bool] = []

        def spy_capture_reading(row, device_name, capture_dir_arg, *, write_excel, write_image=False, scale="kilo"):  # noqa: ANN001
            with locks.get(ENDPOINT).background() as ok:
                acquired_during_capture.append(ok)

        monkeypatch.setattr(billing_module, "capture_reading", spy_capture_reading)

        read_and_store_billing(device_id, locks=locks, now=NOW)

        assert acquired_during_capture == [True]


def _utc(moment: datetime) -> datetime:
    """SQLite hands back a naive datetime for a `DateTime(timezone=True)`
    column — re-attach UTC before comparing against an aware value, same
    rule the API layer's own `_ensure_utc` field validators apply."""
    return moment if moment.tzinfo is not None else moment.replace(tzinfo=UTC)


def _spy_render_billing_png(monkeypatch: pytest.MonkeyPatch) -> list[list[object]]:
    """Patch `capture.service`'s imported `render_billing_png` with a spy
    that records the `rows` argument of every call and still delegates to
    the real renderer, so the write path completes normally. Used to prove
    the D4 row-selection query without inspecting the query directly."""
    import arichds.capture.service as service_module

    captured_rows: list[list[object]] = []
    original = service_module.render_billing_png

    def spy(rows, device_name, scale="kilo"):  # noqa: ANN001
        captured_rows.append(list(rows))
        return original(rows, device_name, scale=scale)

    monkeypatch.setattr(service_module, "render_billing_png", spy)
    return captured_rows


class TestPngIsWrittenWhenImageExportIsOn:
    def test_the_png_is_written_alongside_pdf_and_xlsx(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        license_features(["billing", "auto_capture", "billing_excel_export", "billing_image_export"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        files = captured_files(capture_dir)
        assert any(f.suffix == ".png" for f in files)


class TestT10ImageExportOffWritesNoPng:
    def test_pdf_exists_and_png_does_not(
        self, device_id: int, fake_meter: FakeMeterState, capture_dir: Path, license_features
    ) -> None:
        """T10 (D1/D3) — with `auto_capture` on and `billing_image_export`
        off, the `.pdf` exists and the `.png` does not. Both asserted
        absolutely."""
        license_features(["billing", "auto_capture"])  # no billing_image_export
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        files = captured_files(capture_dir)
        assert any(f.suffix == ".pdf" for f in files)
        assert not any(f.suffix == ".png" for f in files)


class TestT11PngFailureNeverUndoesThePdf:
    def test_a_png_render_failure_is_logged_and_the_pdf_and_read_still_succeed(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """T11 (D3) — a PNG render failure must not undo the PDF, must not
        fail the read, and must be logged."""
        import arichds.capture.service as service_module

        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        fake_meter.billing_rows = [ENTRY_CLOSED]

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("png blew up")

        monkeypatch.setattr(service_module, "render_billing_png", boom)

        with caplog.at_level("ERROR"):
            result = read_and_store_billing(device_id, now=NOW)

        assert result.error is None
        files = captured_files(capture_dir)
        assert any(f.suffix == ".pdf" for f in files)
        assert not any(f.suffix == ".png" for f in files)
        assert any("png" in record.message.lower() for record in caplog.records)


class TestPngRowSelectionD4:
    """T4-T9 (D4/D6, issue #35) — the ten-row window the eager PNG capture
    builds, proven end to end through the real write path by spying on the
    renderer's `rows` argument rather than inspecting the query directly."""

    def test_t4_bounded_to_the_newest_ten_of_thirteen_closed_periods(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from arichds.db.models import BillingReading as BillingReadingRow

        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        pre_existing_dates = [ENTRY_CLOSED.bill_date - timedelta(days=31 * (i + 1)) for i in range(12)]
        with session_scope() as session:
            for bill_date in pre_existing_dates:
                session.add(
                    BillingReadingRow(
                        device_id=device_id,
                        bill_date=bill_date,
                        read_at=NOW,
                        record_status=None,
                        source=SOURCE_DLMS,
                        meter_serial="1232002893",
                    )
                )

        fake_meter.billing_rows = [ENTRY_CLOSED]  # the 13th closed period, the newest of all 13
        read_and_store_billing(device_id, now=NOW)

        assert len(captured_rows) == 1
        rows = captured_rows[0]
        assert len(rows) == 10
        expected_dates = sorted([ENTRY_CLOSED.bill_date, *pre_existing_dates], reverse=True)[:10]
        assert [_utc(row.bill_date) for row in rows] == expected_dates

    def test_t5_bill_date_lte_excludes_a_newer_period_committed_in_the_same_read(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """The image anchored at the OLDER of two periods closed in one read
        must show that older period as its own newest row, and must not
        contain the newer one."""
        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        older = BillingReading(
            bill_date=datetime(2026, 6, 30, 17, 0, 0, tzinfo=UTC),
            source=SOURCE_DLMS,
            is_open=False,
            meter_serial="1232002893",
            import_active_kwh_total=190000.0,
        )
        newer = ENTRY_CLOSED  # 2026-07-31
        fake_meter.billing_rows = [older, newer]

        read_and_store_billing(device_id, now=NOW)

        assert len(captured_rows) == 2
        older_rows, newer_rows = captured_rows
        assert [_utc(row.bill_date) for row in older_rows] == [older.bill_date]
        newer_dates = [_utc(row.bill_date) for row in newer_rows]
        assert newer.bill_date in newer_dates
        assert older.bill_date in newer_dates  # older period is still visible from the newer anchor

    def test_t6_meter_serial_scopes_the_image_to_its_own_serial(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """One `device_id`, two different `meter_serial` values (a meter
        swap) — the image filed under serial A must contain only serial-A
        rows."""
        from arichds.db.models import BillingReading as BillingReadingRow

        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        with session_scope() as session:
            session.add(
                BillingReadingRow(
                    device_id=device_id,
                    bill_date=ENTRY_CLOSED.bill_date - timedelta(days=31),
                    read_at=NOW,
                    record_status=None,
                    source=SOURCE_DLMS,
                    meter_serial="OTHER_SERIAL",
                )
            )

        fake_meter.billing_rows = [ENTRY_CLOSED]  # meter_serial="1232002893"
        read_and_store_billing(device_id, now=NOW)

        assert len(captured_rows) == 1
        rows = captured_rows[0]
        assert all(row.meter_serial == "1232002893" for row in rows)

    def test_t7_the_open_period_never_appears_in_the_image(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        fake_meter.billing_rows = [ENTRY_OPEN, ENTRY_CLOSED]

        read_and_store_billing(device_id, now=NOW)

        assert len(captured_rows) == 1  # only the closed period gets a capture at all
        rows = captured_rows[0]
        assert all(row.record_status is None for row in rows)
        assert ENTRY_OPEN.bill_date not in [_utc(row.bill_date) for row in rows]

    def test_t7_an_open_period_with_the_same_bill_date_as_the_anchor_is_still_excluded(
        self,
        device_id: int,
        capture_dir: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Review round, problem 1 — the case above (`ENTRY_OPEN.bill_date`
        `2026-08-07` vs `ENTRY_CLOSED.bill_date` `2026-07-31`) is already
        excluded by `bill_date <= anchor.bill_date` alone, so it never
        exercises the `record_status IS NULL` clause — deleting that clause
        from `_png_source_rows` left the full suite green. `<=` is
        *inclusive*: an Open Period whose `bill_date` exactly equals the
        closed anchor's is the case only `record_status` can catch. Seeded
        directly through the session (not the read path, which does not
        produce this pairing) and driven straight at `write_png_capture` —
        the assertion under test is the query, not the read path."""
        from arichds.capture.service import write_png_capture
        from arichds.db.models import BillingReading as BillingReadingRow

        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        with session_scope() as session:
            closed = BillingReadingRow(
                device_id=device_id,
                bill_date=ENTRY_CLOSED.bill_date,
                read_at=NOW,
                record_status=None,
                source=SOURCE_DLMS,
                meter_serial="1232002893",
            )
            session.add(closed)
            session.add(
                BillingReadingRow(
                    device_id=device_id,
                    bill_date=ENTRY_CLOSED.bill_date,  # exactly equal — `<=` alone would admit it
                    read_at=NOW,
                    record_status="open",
                    source=SOURCE_DLMS,
                    meter_serial="1232002893",
                )
            )
            session.flush()

            png_target = capture_dir / "1232002893" / "equal-bill-date.png"
            write_png_capture(closed, "Main Incomer", png_target, capture_dir)

        assert len(captured_rows) == 1
        rows = captured_rows[0]
        assert len(rows) == 1  # only the closed row — the Open Period, despite the equal bill_date, is excluded
        assert all(row.record_status is None for row in rows)

    def test_t9_rows_are_ordered_newest_first(
        self,
        device_id: int,
        fake_meter: FakeMeterState,
        capture_dir: Path,
        license_features,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from arichds.db.models import BillingReading as BillingReadingRow

        license_features(["billing", "auto_capture", "billing_image_export"])
        set_capture_dir(capture_dir)
        captured_rows = _spy_render_billing_png(monkeypatch)

        older_dates = [ENTRY_CLOSED.bill_date - timedelta(days=31), ENTRY_CLOSED.bill_date - timedelta(days=62)]
        with session_scope() as session:
            for bill_date in older_dates:
                session.add(
                    BillingReadingRow(
                        device_id=device_id,
                        bill_date=bill_date,
                        read_at=NOW,
                        record_status=None,
                        source=SOURCE_DLMS,
                        meter_serial="1232002893",
                    )
                )

        fake_meter.billing_rows = [ENTRY_CLOSED]
        read_and_store_billing(device_id, now=NOW)

        rows = captured_rows[0]
        dates = [_utc(row.bill_date) for row in rows]
        assert len(set(dates)) >= 3
        assert dates == sorted(dates, reverse=True)
        assert dates[0] == ENTRY_CLOSED.bill_date
        assert dates[-1] == min(older_dates)
