"""``GET /api/billing/captures/{reading_id}`` — render-on-miss download (decision
14, ADR 0010, issue #22).

Existing file -> stream it. Missing -> render, write, serve — this is what
replaces v1's regenerate endpoint, its permanently-failed-capture state, and
its two-second frontend retry. Per-format feature gate on top of the router's
own `billing` gate: `format=pdf` needs `auto_capture`, `format=xlsx` needs
`billing_excel_export`.
"""

from __future__ import annotations

import base64
import io
from datetime import UTC, datetime
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

pytestmark = pytest.mark.usefixtures("fake_meter")

BILL_DATE = datetime(2026, 7, 31, 17, 0, 0, tzinfo=UTC)

#: The smallest possible valid PNG (1x1, transparent) — no Pillow, no
#: browser. `render_billing_png` (ADR 0017, issue #38) drives real headless
#: Edge and needs a real admin account and a real port to navigate to;
#: every HTTP-level test here that only cares about the *hardened write and
#: serve* path around it stubs the renderer with this instead.
_FAKE_PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def stub_render_billing_png(monkeypatch: pytest.MonkeyPatch) -> None:
    """Patch `capture.service`'s imported `render_billing_png` to return
    :data:`_FAKE_PNG_BYTES` — the hardened-write/serve path is still
    exercised end to end; only the real browser drive is skipped."""
    import arichds.capture.service as service_module

    monkeypatch.setattr(service_module, "render_billing_png", lambda *args, **kwargs: _FAKE_PNG_BYTES)


def make_device(admin_client: TestClient) -> int:
    response = admin_client.post(
        "/api/devices",
        json={
            "name": "Main Incomer",
            "brand": "mitsu",
            "model": "smw110",
            "site_name": "Plant A",
            "transport": {"kind": "net", "host": "127.0.0.1", "port": 4059},
            "password": "hunter2",
        },
    )
    assert response.status_code == 201, response.text
    return response.json()["data"]["id"]


def seed_closed(device_id: int, meter_serial: str | None = "1232002893") -> int:
    from arichds.db.models import BillingReading
    from arichds.db.session import session_scope

    with session_scope() as session:
        row = BillingReading(
            device_id=device_id,
            bill_date=BILL_DATE,
            read_at=BILL_DATE,
            record_status=None,
            source="dlms",
            meter_serial=meter_serial,
            import_active_kwh_total=198685.030,
        )
        session.add(row)
        session.flush()
        return row.id


def seed_open(device_id: int) -> int:
    from arichds.db.models import BillingReading
    from arichds.db.session import session_scope

    with session_scope() as session:
        row = BillingReading(
            device_id=device_id,
            bill_date=BILL_DATE,
            read_at=BILL_DATE,
            record_status="open",
            source="dlms",
            meter_serial="1232002893",
        )
        session.add(row)
        session.flush()
        return row.id


def set_capture_dir(admin_client: TestClient, path: Path) -> None:
    response = admin_client.put("/api/billing/settings", json={"capture_dir": str(path)})
    assert response.status_code == 200, response.text


class TestRenderOnMiss:
    def test_a_missing_pdf_is_rendered_written_and_served(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "pdf"})

        assert response.status_code == 200, response.text
        assert response.content.startswith(b"%PDF")
        assert response.headers["content-type"] == "application/pdf"
        written = list(capture_dir.rglob("*.pdf"))
        assert len(written) == 1

    def test_xlsx_format_is_rendered_written_and_served(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "xlsx"})

        assert response.status_code == 200, response.text
        assert len(response.content) > 0

    def test_default_format_is_pdf(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}")

        assert response.status_code == 200, response.text
        assert response.content.startswith(b"%PDF")


class TestDisplayUnitScaleReachesTheDownload:
    def test_base_scale_is_applied_to_a_render_on_miss_xlsx(self, admin_client: TestClient, tmp_path: Path) -> None:
        from openpyxl import load_workbook

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        put_response = admin_client.put("/api/settings/display", json={"display_unit_scale": "base"})
        assert put_response.status_code == 200, put_response.text
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "xlsx"})

        assert response.status_code == 200, response.text
        values = [
            cell.value
            for row_cells in load_workbook(io.BytesIO(response.content)).active.iter_rows()
            for cell in row_cells
        ]
        assert "Import Active Wh Total" in values
        assert "Import Active kWh Total" not in values


class TestExistingFileIsServedWithoutRerendering:
    def test_a_second_download_does_not_call_the_renderer_again(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        import arichds.capture.pdf as pdf_module

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        first = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "pdf"})
        assert first.status_code == 200, first.text

        calls = []
        original = pdf_module.render_billing_pdf

        def spy(*args, **kwargs):  # noqa: ANN002, ANN003, ANN201
            calls.append(1)
            return original(*args, **kwargs)

        monkeypatch.setattr(pdf_module, "render_billing_pdf", spy)

        second = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "pdf"})

        assert second.status_code == 200, second.text
        assert second.content == first.content
        assert calls == []  # the renderer was never invoked the second time


class TestOpenPeriodHasNoCapture:
    def test_downloading_the_open_period_is_404(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_open(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}")

        assert response.status_code == 404, response.text


class TestUnknownId:
    def test_unknown_reading_id_is_404(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)

        response = admin_client.get("/api/billing/captures/999999")

        assert response.status_code == 404, response.text


class TestCaptureDirNotConfigured:
    def test_download_is_404_when_capture_dir_is_unset(self, admin_client: TestClient) -> None:
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        response = admin_client.get(f"/api/billing/captures/{reading_id}")

        assert response.status_code == 404, response.text


class TestWriteFailureIsA500NotAnUnhandledCrash:
    def test_an_oserror_during_write_becomes_a_structured_500(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """A disk-full or permission error inside the hardened write is a real
        possibility on a network-share capture_dir — it must come back as a
        clean 500 with a message, never an unhandled 500 with no detail."""
        import arichds.api.billing as billing_module

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        def boom(*args: object, **kwargs: object) -> None:
            raise OSError("disk full")

        monkeypatch.setattr(billing_module, "write_pdf_capture", boom)

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "pdf"})

        assert response.status_code == 500, response.text
        assert "disk full" in response.json()["detail"]


class TestNullMeterSerialIs422NotACrash:
    def test_a_row_with_no_meter_serial_is_422_on_download(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id, meter_serial=None)

        response = admin_client.get(f"/api/billing/captures/{reading_id}")

        assert response.status_code == 422, response.text


class TestBillingImageEndpoint:
    """``GET /api/billing/captures/image?device_id={id}`` — device-keyed,
    render-on-miss (D11, ADR 0014/0015, issue #35)."""

    def test_t14_a_missing_png_is_rendered_written_and_served(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """T14 — both halves in one test: the response body is PNG bytes,
        and the file now exists under capture_dir/<serial>/<stem>.png."""
        stub_render_billing_png(monkeypatch)
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 200, response.text
        assert response.content.startswith(b"\x89PNG\r\n\x1a\n")
        assert response.headers["content-type"] == "image/png"
        written = list((capture_dir / "1232002893").glob("*.png"))
        assert len(written) == 1

    def test_t15_the_filename_stem_is_the_newest_closed_periods(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """T15 — with several closed periods, the endpoint's derived filename
        stem is the NEWEST closed period's, not the oldest or an arbitrary
        one."""
        from arichds.db.models import BillingReading
        from arichds.db.session import session_scope

        stub_render_billing_png(monkeypatch)
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)  # 2026-07-31
        with session_scope() as session:
            session.add(
                BillingReading(
                    device_id=device_id,
                    bill_date=datetime(2026, 6, 30, 17, 0, 0, tzinfo=UTC),
                    read_at=BILL_DATE,
                    record_status=None,
                    source="dlms",
                    meter_serial="1232002893",
                )
            )

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 200, response.text
        assert "content-disposition" in response.headers
        assert "2026-07-31" in response.headers["content-disposition"]
        assert "2026-06-30" not in response.headers["content-disposition"]

    def test_t16_zero_closed_periods_is_404_not_a_500(self, admin_client: TestClient, tmp_path: Path) -> None:
        """T16 — a device with zero closed periods must be a clean 404 with
        the specific sentence, not a 500."""
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 404, response.text
        assert response.json()["detail"] == "This device has no closed billing period yet."

    def test_only_open_period_is_also_a_404(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_open(device_id)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 404, response.text

    def test_capture_dir_unconfigured_is_404(self, admin_client: TestClient) -> None:
        device_id = make_device(admin_client)
        seed_closed(device_id)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 404, response.text

    def test_a_row_with_no_meter_serial_is_422(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id, meter_serial=None)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 422, response.text

    def test_a_second_request_does_not_re_render(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        import arichds.capture.service as service_module

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)

        calls: list[int] = []

        def spy(*args, **kwargs):  # noqa: ANN002, ANN003, ANN201
            calls.append(1)
            return _FAKE_PNG_BYTES

        monkeypatch.setattr(service_module, "render_billing_png", spy)

        first = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})
        assert first.status_code == 200, first.text
        assert calls == [1]  # the first request DOES render — proves the spy itself is live

        second = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert second.status_code == 200, second.text
        assert second.content == first.content
        assert calls == [1]  # unchanged — the renderer was never invoked the second time

    def test_t13_requires_billing_image_export(self, admin_client: TestClient, tmp_path: Path, relicense) -> None:
        """T13 — a licence with `billing` but not `billing_image_export` gets
        `FEATURE_DISABLED` with `reason == "billing_image_export"`."""
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)

        relicense(admin_client, features=["billing"])  # no billing_image_export

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 403, response.text
        assert response.json()["error"]["code"] == "FEATURE_DISABLED"
        assert response.json()["error"]["reason"] == "billing_image_export"

    def test_a_write_failure_is_a_500_not_a_crash(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        import arichds.api.billing as billing_module

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)

        def boom(*args: object, **kwargs: object) -> None:
            raise OSError("disk full")

        monkeypatch.setattr(billing_module, "write_png_capture", boom)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 500, response.text
        assert "disk full" in response.json()["detail"]

    def test_a_browsercaptureerror_is_a_500_naming_the_browser_not_an_unhandled_crash(
        self, admin_client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """`BrowserCaptureError` (ADR 0017, issue #38) is a plain
        `Exception`, not an `OSError` — today's `except OSError` at
        `download_billing_image` does not catch it, so without an explicit
        clause this would be an unhandled 500 with no detail."""
        import arichds.api.billing as billing_module
        from arichds.capture.screenshot import BrowserCaptureError

        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        seed_closed(device_id)

        def boom(*args: object, **kwargs: object) -> None:
            raise BrowserCaptureError("Microsoft Edge was not found on this machine.")

        monkeypatch.setattr(billing_module, "write_png_capture", boom)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": device_id})

        assert response.status_code == 500, response.text
        assert "Microsoft Edge" in response.json()["detail"]

    def test_an_unknown_device_id_is_404(self, admin_client: TestClient, tmp_path: Path) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)

        response = admin_client.get("/api/billing/captures/image", params={"device_id": 999999})

        assert response.status_code == 404, response.text


class TestPerFormatFeatureGate:
    def test_pdf_format_requires_auto_capture(self, admin_client: TestClient, tmp_path: Path, relicense) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        relicense(admin_client, features=["billing"])  # no auto_capture, no billing_excel_export

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "pdf"})

        assert response.status_code == 403, response.text
        assert response.json()["error"]["code"] == "FEATURE_DISABLED"
        assert response.json()["error"]["reason"] == "auto_capture"

    def test_xlsx_format_requires_billing_excel_export(
        self, admin_client: TestClient, tmp_path: Path, relicense
    ) -> None:
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        set_capture_dir(admin_client, capture_dir)
        device_id = make_device(admin_client)
        reading_id = seed_closed(device_id)

        relicense(admin_client, features=["billing", "auto_capture"])  # no billing_excel_export

        response = admin_client.get(f"/api/billing/captures/{reading_id}", params={"format": "xlsx"})

        assert response.status_code == 403, response.text
        assert response.json()["error"]["reason"] == "billing_excel_export"
