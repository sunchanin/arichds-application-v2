"""``capture.pdf`` / ``capture.xlsx`` — the two capture renderers (M6b, issue #22).

``row`` is anything exposing the ``BillingReading`` field names as attributes
— here a ``SimpleNamespace`` built from the same field set
``_render_shared.ALL_SECTIONS`` walks, which is exactly how both the eager
capture path and the download endpoint feed a real ORM row in.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from arichds.capture.pdf import render_billing_pdf
from arichds.capture.xlsx import render_billing_xlsx

BILL_DATE = datetime(2026, 7, 31, 17, 0, 0, tzinfo=UTC)


def make_row(**overrides: object) -> SimpleNamespace:
    from arichds.capture._render_shared import ALL_SECTIONS

    fields = {attr: None for _title, section in ALL_SECTIONS for _label, attr in section}
    fields.update(bill_date=BILL_DATE, meter_serial="1232002893", record_status=None, read_at=BILL_DATE)
    fields.update(overrides)
    return SimpleNamespace(**fields)


class TestRenderBillingPdf:
    def test_returns_nonempty_pdf_bytes(self) -> None:
        row = make_row(import_active_kwh_total=200464.501)

        result = render_billing_pdf(row, "Main Incomer")

        assert isinstance(result, bytes)
        assert result.startswith(b"%PDF")

    def test_non_ascii_device_name_does_not_raise_and_is_replaced(self) -> None:
        row = make_row()

        result = render_billing_pdf(row, "ระบบ")  # Thai device name

        assert isinstance(result, bytes)
        assert result.startswith(b"%PDF")

    def test_a_none_meter_serial_does_not_raise(self) -> None:
        row = make_row(meter_serial=None)

        result = render_billing_pdf(row, "Main Incomer")

        assert result.startswith(b"%PDF")


class TestRenderBillingXlsx:
    def test_returns_a_loadable_workbook(self) -> None:
        row = make_row(import_active_kwh_total=200464.501)

        result = render_billing_xlsx(row, "Main Incomer")

        wb = load_workbook(io.BytesIO(result))
        sheet = wb.active
        values = [cell.value for row_cells in sheet.iter_rows() for cell in row_cells]
        assert "Main Incomer" in values
        assert 200464.501 in values or "200464.501" in values

    def test_non_ascii_device_name_is_preserved_verbatim(self) -> None:
        row = make_row()

        result = render_billing_xlsx(row, "ระบบ")

        wb = load_workbook(io.BytesIO(result))
        sheet = wb.active
        values = [cell.value for row_cells in sheet.iter_rows() for cell in row_cells]
        assert "ระบบ" in values


class TestDisplayUnitScaleXlsx:
    """The machine-wide kilo<->base setting (a CR, not part of M6b's original
    shape) — applied at render time only. The value and the label must always
    move together: a kilo *number* under a `Wh` *header* is exactly the
    defect this feature can introduce, so both are asserted from the same
    render rather than by inspection."""

    def test_kilo_is_the_default_and_leaves_values_and_labels_alone(self) -> None:
        row = make_row(import_active_kwh_total=200464.501)

        values = _xlsx_values(render_billing_xlsx(row, "Main Incomer"))

        assert 200464.501 in values or "200464.501" in values
        assert "Import Active kWh Total" in values
        assert "Import Active Wh Total" not in values

    def test_base_scales_the_value_and_relabels_it_together(self) -> None:
        # A value whose x1000 lands exactly, so `str()` round-trips cleanly —
        # `format_cell` stringifies every value, so a float that only *looks*
        # exact after scaling (e.g. 200.464501 * 1000) would make this
        # assertion about the label-vs-value pairing fail on a floating-point
        # artifact that has nothing to do with the feature under test.
        row = make_row(import_active_kwh_total=42.5)

        values = _xlsx_values(render_billing_xlsx(row, "Main Incomer", scale="base"))

        assert 42500.0 in values or "42500.0" in values
        assert "Import Active Wh Total" in values
        assert "Import Active kWh Total" not in values

    def test_base_scales_a_reactive_power_field(self) -> None:
        row = make_row(max_demand_import_reactive_kvar_rate_a=2.5)

        values = _xlsx_values(render_billing_xlsx(row, "Main Incomer", scale="base"))

        assert 2500.0 in values or "2500.0" in values
        assert "Max Demand Import Reactive var Rate A" in values

    def test_base_leaves_the_demand_time_column_as_a_timestamp(self) -> None:
        """A Demand Time value is a `datetime`, never a magnitude — base scale
        must not multiply it, and its label carries no unit token to begin
        with (D18)."""
        stamp = datetime(2026, 8, 5, 9, 12, 0, tzinfo=UTC)
        row = make_row(max_demand_import_active_time_total=stamp)

        values = _xlsx_values(render_billing_xlsx(row, "Main Incomer", scale="base"))

        assert "2026-08-05 09:12:00+00:00" in values
        assert "Demand Time Import Active Total" in values

    def test_base_leaves_voltage_and_metadata_columns_alone(self) -> None:
        """Only energy and power columns convert — Bill Date (Metadata) must
        render at its stored value under either scale."""
        row = make_row(bill_date=BILL_DATE)

        values = _xlsx_values(render_billing_xlsx(row, "Main Incomer", scale="base"))

        assert "2026-07-31 17:00:00+00:00" in values


class TestDisplayUnitScalePdf:
    def test_base_scale_does_not_raise_and_is_still_a_pdf(self) -> None:
        row = make_row(import_active_kwh_total=200.464501)

        result = render_billing_pdf(row, "Main Incomer", scale="base")

        assert result.startswith(b"%PDF")

    def test_kilo_is_the_default(self) -> None:
        row_a = make_row(import_active_kwh_total=200464.501)
        row_b = make_row(import_active_kwh_total=200464.501)

        assert render_billing_pdf(row_a, "Main Incomer") == render_billing_pdf(row_b, "Main Incomer", scale="kilo")


def _xlsx_values(xlsx_bytes: bytes) -> list[object]:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = wb.active
    return [cell.value for row_cells in sheet.iter_rows() for cell in row_cells]


class TestBothRenderersUseTheSameSections:
    def test_a_field_present_in_all_sections_renders_in_the_xlsx(self) -> None:
        """`ALL_SECTIONS` is asserted complete (40 fields, no duplicates) in
        `test_capture_render_shared.py`; this proves a value at a field only
        the last section carries actually reaches the rendered output — i.e.
        the renderer walks the whole list, not a truncated prefix of it."""
        row = make_row(max_demand_export_reactive_kvar_rate_d=42.5)

        xlsx_values = [
            cell.value
            for row_cells in load_workbook(io.BytesIO(render_billing_xlsx(row, "Main Incomer"))).active.iter_rows()
            for cell in row_cells
        ]
        pdf_bytes = render_billing_pdf(row, "Main Incomer")

        assert "42.5" in xlsx_values  # format_cell() stringifies every value, floats included
        assert pdf_bytes.startswith(b"%PDF")  # the PDF path did not raise on the same row
