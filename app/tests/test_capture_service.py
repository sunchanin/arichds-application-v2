"""``capture.service`` — wires the renderers and the hardened write together
(M6b, issue #22). The one place both callers (the eager write in
``acquisition/billing.py`` and the render-on-miss download in
``api/billing.py``) build a path and write a file, so they cannot drift.
"""

from __future__ import annotations

import base64
from datetime import UTC, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest

from arichds.capture.service import capture_reading, capture_target_paths, write_png_capture

BILL_DATE = datetime(2026, 7, 31, 17, 0, 0, tzinfo=UTC)

#: The smallest possible valid PNG (1x1, transparent) — no Pillow, no
#: browser. `render_billing_png` (ADR 0017, issue #38) needs a real Edge
#: install and a real database to run at all, so every test here that only
#: cares about the *hardened write* around it stubs the renderer with this
#: instead of exercising the real headless-screenshot path.
FAKE_PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def stub_render_billing_png(monkeypatch: pytest.MonkeyPatch) -> None:
    """Patch `capture.service`'s imported `render_billing_png` to return
    :data:`FAKE_PNG_BYTES` — the hardened-write path (target/allowlist/
    fsync) is still exercised end to end; only the actual browser drive is
    skipped."""
    import arichds.capture.service as service_module

    monkeypatch.setattr(service_module, "render_billing_png", lambda *args, **kwargs: FAKE_PNG_BYTES)


def make_row(**overrides: object) -> SimpleNamespace:
    from arichds.capture._render_shared import ALL_SECTIONS

    fields = {attr: None for _title, section in ALL_SECTIONS for _label, attr in section}
    fields.update(bill_date=BILL_DATE, meter_serial="1232002893", record_status=None, read_at=BILL_DATE)
    fields.update(overrides)
    return SimpleNamespace(**fields)


class TestCaptureTargetPaths:
    def test_the_path_follows_the_fixed_convention(self, tmp_path: Path) -> None:
        pdf_path, xlsx_path, png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)

        assert pdf_path == (tmp_path / "1232002893" / "2026-07-31_170000.pdf").resolve()
        assert xlsx_path == (tmp_path / "1232002893" / "2026-07-31_170000.xlsx").resolve()
        assert png_path == (tmp_path / "1232002893" / "2026-07-31_170000.png").resolve()

    def test_a_naive_bill_date_is_treated_as_utc(self, tmp_path: Path) -> None:
        naive = BILL_DATE.replace(tzinfo=None)
        pdf_path, _xlsx_path, _png_path = capture_target_paths(tmp_path, "1232002893", naive)

        assert pdf_path.name == "2026-07-31_170000.pdf"

    def test_an_unsafe_serial_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError):
            capture_target_paths(tmp_path, "../escape", BILL_DATE)

    def test_t12_all_three_paths_share_one_stem_with_the_right_suffixes(self, tmp_path: Path) -> None:
        """T12 (D2/ADR 0015): the three formats share **one** filename stem —
        only the extension differs. Asserted directly, not just inferred from
        the fixed-convention test above."""
        pdf_path, xlsx_path, png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)

        assert pdf_path.stem == xlsx_path.stem == png_path.stem
        assert (pdf_path.suffix, xlsx_path.suffix, png_path.suffix) == (".pdf", ".xlsx", ".png")


class TestCaptureReading:
    def test_writes_only_the_pdf_when_excel_is_off(self, tmp_path: Path) -> None:
        row = make_row()

        capture_reading(row, "Main Incomer", tmp_path, write_excel=False, write_image=False)

        pdf_path, xlsx_path, _png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert not xlsx_path.exists()

    def test_writes_both_when_excel_is_on(self, tmp_path: Path) -> None:
        row = make_row()

        capture_reading(row, "Main Incomer", tmp_path, write_excel=True, write_image=False)

        pdf_path, xlsx_path, _png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert xlsx_path.exists()

    def test_a_missing_meter_serial_writes_nothing_and_does_not_raise(self, tmp_path: Path) -> None:
        row = make_row(meter_serial=None)

        capture_reading(row, "Main Incomer", tmp_path, write_excel=True, write_image=False)  # must not raise

        assert list(tmp_path.iterdir()) == []

    def test_the_scale_argument_reaches_the_xlsx_render(self, tmp_path: Path) -> None:
        """`capture_reading`'s `scale` must reach the renderer, not just be
        accepted and dropped — proven by reading the label back off the
        written file rather than by inspecting the call."""
        from openpyxl import load_workbook

        row = make_row(import_active_kwh_total=42.5)

        capture_reading(row, "Main Incomer", tmp_path, write_excel=True, write_image=False, scale="base")

        _pdf_path, xlsx_path, _png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        values = [cell.value for sheet_row in load_workbook(xlsx_path).active.iter_rows() for cell in sheet_row]
        assert "Import Active Wh Total" in values
        assert 42500.0 in values or "42500.0" in values

    def test_an_xlsx_write_failure_does_not_undo_the_pdf(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        import arichds.capture.service as service_module

        row = make_row()

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("xlsx blew up")

        monkeypatch.setattr(service_module, "write_xlsx_capture", boom)

        capture_reading(row, "Main Incomer", tmp_path, write_excel=True, write_image=False)  # must not raise

        pdf_path, xlsx_path, _png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert not xlsx_path.exists()

    def test_writes_the_png_when_write_image_is_on(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        stub_render_billing_png(monkeypatch)
        row = make_row()

        capture_reading(row, "Main Incomer", tmp_path, write_excel=False, write_image=True)

        pdf_path, _xlsx_path, png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert png_path.exists()
        assert png_path.read_bytes() == FAKE_PNG_BYTES

    def test_t10_write_image_off_never_writes_a_png(self, tmp_path: Path) -> None:
        """T10 (D1/D3) — with `write_image` off, the PDF is still written and
        the PNG is not, absolutely: both are asserted, not just the PNG's
        absence."""
        row = make_row()

        capture_reading(row, "Main Incomer", tmp_path, write_excel=True, write_image=False)

        pdf_path, xlsx_path, png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert xlsx_path.exists()
        assert not png_path.exists()

    def test_t11_a_png_write_failure_does_not_undo_the_pdf_and_is_logged(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        """T11 (D3) — the PNG write is a swallowed side effect exactly like
        the xlsx write: a failure must not undo the PDF and must not raise."""
        import arichds.capture.service as service_module

        row = make_row()

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("png blew up")

        monkeypatch.setattr(service_module, "write_png_capture", boom)

        with caplog.at_level("ERROR"):
            capture_reading(row, "Main Incomer", tmp_path, write_excel=False, write_image=True)  # must not raise

        pdf_path, _xlsx_path, png_path = capture_target_paths(tmp_path, "1232002893", BILL_DATE)
        assert pdf_path.exists()
        assert not png_path.exists()
        assert any("png" in record.message.lower() for record in caplog.records)


class TestWritePngCaptureDetachedRow:
    """Review round, problem 4 — a *mapped* ``BillingReading`` with no
    attached session (e.g. detached/transient) is a different case from a
    ``SimpleNamespace`` test row (unmapped — the documented, silent
    fallback). Rendering it as a single row is still correct (there is no
    session to run the D4 query against), but it must never be silent: a
    file named for ten periods that quietly holds one, with nothing on disk
    or in the log, is exactly the failure ADR 0015 warns about."""

    def test_a_detached_mapped_row_warns_and_still_renders_the_single_row(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        import arichds.capture.service as service_module
        from arichds.db.models import BillingReading as BillingReadingRow

        # Constructed, never added to any Session — a *mapped* instance
        # (unlike a SimpleNamespace) that `object_session()` resolves to
        # `None` rather than raising `UnmappedInstanceError` on.
        row = BillingReadingRow(
            device_id=1,
            bill_date=BILL_DATE,
            read_at=BILL_DATE,
            record_status=None,
            source="dlms",
            meter_serial="1232002893",
        )

        captured_rows: list[list[object]] = []

        def spy(rows, device_name):  # noqa: ANN001
            captured_rows.append(list(rows))
            return FAKE_PNG_BYTES

        monkeypatch.setattr(service_module, "render_billing_png", spy)

        with caplog.at_level("WARNING"):
            write_png_capture(row, "Main Incomer", tmp_path / "1232002893" / "detached.png", tmp_path)

        assert len(captured_rows) == 1
        assert captured_rows[0] == [row]  # still renders — just the one row it was given
        assert any("session" in record.message.lower() for record in caplog.records)

    def test_an_unmapped_simplenamespace_row_stays_silent(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        """The documented, intended fallback (pdf.py/xlsx.py's own row
        shape) — a plain test double, not an ORM row — must NOT warn; this
        is the case the review explicitly asked to keep silent."""
        stub_render_billing_png(monkeypatch)
        row = make_row()

        with caplog.at_level("WARNING"):
            write_png_capture(row, "Main Incomer", tmp_path / "1232002893" / "namespace.png", tmp_path)

        assert not any("session" in record.message.lower() for record in caplog.records)
